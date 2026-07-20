import streamlit as st
import asyncio
import pandas as pd
import os
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from agents import TechnicalAgent, MomentumAgent, FundamentalAgent, QualityAgent, DeepSentimentAgent, MacroSectorAgent, MicrostructureAgent
from utils import get_data, log_trade

# --- UI Configuration ---
st.set_page_config(page_title="Tango Pro-Screener", layout="wide")

# High-Contrast Dark Theme CSS
st.markdown("""
    <style>
    .stApp { background-color: #0F172A !important; color: #F8FAFC !important; }
    h1 { color: #38BDF8 !important; font-family: 'Helvetica', sans-serif; font-weight: 700; }
    h2, h3, h4, h5, h6 { color: #94A3B8 !important; }
    p, span, label { color: #E2E8F0 !important; }
    
    div[data-testid="metric-container"] {
        background-color: #1E293B !important; 
        padding: 18px; border-radius: 12px; border: 1px solid #334155 !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.4);
    }
    div[data-testid="metric-container"] label { color: #94A3B8 !important; font-size: 0.9rem !important; }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] { color: #F8FAFC !important; font-weight: bold !important; }
    
    div[data-baseweb="select"] > div, div[data-baseweb="input"] > div {
        background-color: #1E293B !important; color: #F8FAFC !important; border-color: #334155 !important;
    }
    div[role="option"] { background-color: #1E293B !important; color: #F8FAFC !important; }
    .stCheckbox label { color: #F1F5F9 !important; }
    
    div[data-testid="stDataFrame"] { background-color: #1E293B !important; border: 1px solid #334155 !important; border-radius: 8px; padding: 8px; }
    
    .stButton > button {
        background-color: #0284C7 !important; color: #FFFFFF !important; border: none !important;
        font-weight: bold !important; border-radius: 6px !important; padding: 8px 16px !important;
    }
    .stButton > button:hover { background-color: #0369A1 !important; color: #FFFFFF !important; }
    </style>
    """, unsafe_allow_html=True)

st.title("📊 Tango Pro-Screener")
st.markdown("### Institutional Quantitative Cluster Framework (7-Factor)")

def export_to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Factor Breakdown"
    header_fill = PatternFill(start_color="2B3A42", end_color="2B3A42", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 80
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- Ticker Loading ---
col1, col2 = st.columns([1, 2])
with col1:
    uploaded_file = st.file_uploader("Upload Ticker CSV", type=['csv'])
    append_ns = st.checkbox("Append '.NS' (NSE Stocks)", value=True)
    
    tickers_list = ["AAPL", "MSFT", "GOOGL"]
    
    if uploaded_file is not None:
        df_upload = pd.read_csv(uploaded_file)
        col_name = 'Symbol' if 'Symbol' in df_upload.columns else df_upload.columns[0]
        raw = [str(t).strip().upper() for t in df_upload[col_name].dropna().tolist()]
        tickers_list = [f"{t}.NS" if append_ns and not t.endswith('.NS') else t for t in raw]
    
    elif os.path.exists("ind_nifty500list.csv"):
        df_default = pd.read_csv("ind_nifty500list.csv")
        tickers_list = [f"{str(t).strip().upper()}.NS" for t in df_default['Symbol'].dropna().tolist()]
    
    selected_ticker = st.selectbox("Select Asset:", options=tickers_list)
    analyze_btn = st.button("Run Quantitative Deep-Scan", type="primary")

# --- Analysis Pipeline ---
with col2:
    if analyze_btn:
        with st.spinner(f"Evaluating 7-Factor metrics for {selected_ticker}..."):
            try:
                df, vix, info, news_list, pcr, social_score, sector_df, sector_symbol, yield_val, crude_val = asyncio.run(get_data(selected_ticker))
                
                sector_name = info.get('sector', 'Unknown') if isinstance(info, dict) else 'Unknown'
                
                # 7-Factor Agents
                t_sig, t_conf, t_res, current_price, current_atr = TechnicalAgent().analyze(df)
                mom_sig, mom_conf, mom_res = MomentumAgent().analyze(df)
                f_sig, f_conf, f_res = FundamentalAgent().analyze(info)
                q_sig, q_conf, q_res = QualityAgent().analyze(info)
                micro_sig, micro_conf, micro_res = MicrostructureAgent().analyze(df, info)
                s_sig, s_conf, s_res = DeepSentimentAgent().analyze(news_list, pcr, social_score)
                m_sig, m_conf, m_res = MacroSectorAgent().analyze(df, sector_df, sector_symbol, yield_val, crude_val, vix, sector_name)
                
                # Decision Resolution Logic
                def map_signal(sig):
                    if sig in ["BUY", "BULLISH"]: return 1
                    if sig in ["SELL", "BEARISH"]: return -1
                    return 0 
                
                score = (map_signal(t_sig) * 0.15 + 
                         map_signal(mom_sig) * 0.15 + 
                         map_signal(m_sig) * 0.15 + 
                         map_signal(f_sig) * 0.15 + 
                         map_signal(q_sig) * 0.15 + 
                         map_signal(micro_sig) * 0.15 +
                         map_signal(s_sig) * 0.10)
                         
                decision = "BUY" if score > 0.25 else "SELL" if score < -0.25 else "HOLD"
                
                # Helper function for grid colorization
                def colorize(sig):
                    if sig in ["BUY", "BULLISH"]: return f"🟢 {sig}"
                    if sig in ["SELL", "BEARISH"]: return f"🔴 {sig}"
                    if sig == "HALT": return f"🛑 {sig}"
                    return f"⚪ {sig}"

                # Calculate the Risk Meter pointer position (Converts -1.0 to 1.0 into a 0% to 100% scale)
                pointer_pct = max(0, min(100, ((score + 1.0) / 2.0) * 100))
                
                # --- DYNAMIC HTML RISK METER ---
                dial_html = f"""
                <div style="margin-bottom: 25px; background-color: #1E293B; padding: 25px; border-radius: 12px; border: 1px solid #334155; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.4);">
                    <h4 style="color: #F8FAFC; margin-top: 0; margin-bottom: 25px; font-size: 1.1rem; text-align: center;">Composite Score Risk Meter</h4>
                    
                    <div style="position: relative; width: 100%; height: 25px; background: linear-gradient(to right, #EF4444 0%, #F59E0B 50%, #10B981 100%); border-radius: 15px;">
                        <!-- Dynamic Indicator Marker -->
                        <div style="position: absolute; top: -15px; left: calc({pointer_pct}% - 12px); width: 24px; display: flex; flex-direction: column; align-items: center;">
                            <div style="width: 0; height: 0; border-left: 12px solid transparent; border-right: 12px solid transparent; border-top: 18px solid #FFFFFF; filter: drop-shadow(0px 2px 2px rgba(0,0,0,0.5));"></div>
                        </div>
                    </div>
                    
                    <div style="display: flex; justify-content: space-between; margin-top: 20px; color: #94A3B8; font-size: 0.85rem; font-weight: 700;">
                        <span style="color: #EF4444;">-1.0 (Heavy Dist/Bearish)</span>
                        <span style="color: #F59E0B;">0.0 (Neutral/Fair Value)</span>
                        <span style="color: #10B981;">+1.0 (Heavy Acc/Bullish)</span>
                    </div>
                    
                    <div style="text-align: center; margin-top: 15px; font-size: 1.5rem; color: #38BDF8; font-weight: 800;">
                        Current Score: {score:.2f}
                    </div>
                </div>
                """
                st.markdown(dial_html, unsafe_allow_html=True)
                
                # Top Metrics Display
                m1, m2, m3 = st.columns(3)
                m1.metric("Ticker", selected_ticker)
                
                if decision == "BUY":
                    m2.metric("Decision", colorize(decision), delta="Positive Expectancy", delta_color="normal")
                elif decision == "SELL":
                    m2.metric("Decision", colorize(decision), delta="Negative Risk", delta_color="inverse")
                else:
                    m2.metric("Decision", colorize(decision), delta="Flat Regime", delta_color="off")
                    
                m3.metric("Entry Price", f"₹{round(current_price, 2)}" if current_price > 0 else "N/A")
                
                # --- COLORIZED ROW-BASED PARAMETER GRID ---
                rows = [
                    {"Factor / Parameter": "Trade Execution", "Signal": colorize(decision), "Detailed Analysis": f"Stop Loss: ₹{current_price - (2*current_atr):.2f}" if current_atr > 0 else "N/A"},
                    {"Factor / Parameter": "1. Technical (RSI)", "Signal": colorize(t_sig), "Detailed Analysis": t_res},
                    {"Factor / Parameter": "2. Momentum (1M/3M)", "Signal": colorize(mom_sig), "Detailed Analysis": mom_res},
                    {"Factor / Parameter": "3. Value (P/E, P/B)", "Signal": colorize(f_sig), "Detailed Analysis": f_res},
                    {"Factor / Parameter": "4. Quality (ROE, D/E)", "Signal": colorize(q_sig), "Detailed Analysis": q_res},
                    {"Factor / Parameter": "5. Flow & Liquidity", "Signal": colorize(micro_sig), "Detailed Analysis": micro_res},
                    {"Factor / Parameter": "6. Macro & Correlation", "Signal": colorize(m_sig), "Detailed Analysis": m_res},
                    {"Factor / Parameter": "7. Sentiment (FinBERT)", "Signal": colorize(s_sig), "Detailed Analysis": s_res},
                ]
                
                results_df = pd.DataFrame(rows)
                
                # Render Row Grid (Using the updated syntax to clear the warning)
                st.dataframe(
                    results_df, 
                    hide_index=True,
                    width="stretch",
                    column_config={
                        "Factor / Parameter": st.column_config.TextColumn("Factor / Parameter", width="medium"),
                        "Signal": st.column_config.TextColumn("Signal", width="small"),
                        "Detailed Analysis": st.column_config.TextColumn("Detailed Analysis", width="large")
                    }
                )
                
                st.download_button("📥 Export Report", data=export_to_excel(results_df), file_name=f"{selected_ticker}_report.xlsx")
                
            except Exception as e:
                st.error(f"Analysis failed: {e}")